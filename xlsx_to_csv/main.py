#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
xlsx_to_csv
递归搜索目录下所有 Excel (.xlsx)，转换为 CSV。

默认行为：
- 每个工作表输出一个 CSV。
- 若只有一个工作表，输出文件名为 <stem>.csv。
- 若有多个工作表，输出文件名为 <stem>__<sheet>.csv。

用法示例：
  python xlsx_to_csv/main.py
  python xlsx_to_csv/main.py --root /path/to/dir
  python xlsx_to_csv/main.py --output-dir ./converted
  python xlsx_to_csv/main.py --sheet "Sheet1"
  python xlsx_to_csv/main.py --workers 4
  python xlsx_to_csv/main.py --force
  python xlsx_to_csv/main.py --dry-run
  python xlsx_to_csv/main.py --exclude .git node_modules dist
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Optional, Set, Tuple

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


@dataclass(frozen=True)
class JobResult:
    xlsx_path: Path
    output_paths: List[Path]
    ok: bool
    elapsed_s: float
    message: str
    created: int
    skipped: int


def _human_rel(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except Exception:
        return str(path)


def _is_excluded(path: Path, exclude_names: Set[str]) -> bool:
    return any(part in exclude_names for part in path.parts)


def _sanitize_filename(name: str) -> str:
    invalid = {'/', '\\', ':', '*', '?', '"', '<', '>', '|'}
    out = "".join("_" if ch in invalid else ch for ch in name)
    return out.strip() or "sheet"


def find_xlsx_files(root: Path, exclude_names: Set[str], include_hidden: bool) -> List[Path]:
    xlsx_files: List[Path] = []
    for p in root.rglob("*.xlsx"):
        if not p.is_file():
            continue
        if _is_excluded(p, exclude_names):
            continue
        if not include_hidden and any(part.startswith(".") for part in p.relative_to(root).parts):
            continue
        xlsx_files.append(p)
    xlsx_files.sort(key=lambda x: str(x).lower())
    return xlsx_files


def _resolve_output_dir(xlsx_path: Path, output_dir: Optional[str]) -> Path:
    if not output_dir:
        return xlsx_path.parent
    out_dir = Path(output_dir).expanduser()
    if not out_dir.is_absolute():
        out_dir = (xlsx_path.parent / out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def _iter_rows_as_strings(ws) -> Iterable[List[str]]:
    for row in ws.iter_rows(values_only=True):
        yield ["" if v is None else str(v) for v in row]


def _write_sheet_csv(ws, csv_path: Path) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        for row in _iter_rows_as_strings(ws):
            writer.writerow(row)


def convert_one(
    xlsx_path: Path,
    root: Path,
    output_dir: Optional[str],
    sheet_name: Optional[str],
    force: bool,
    dry_run: bool,
) -> JobResult:
    t0 = time.time()
    if load_workbook is None:
        return JobResult(xlsx_path, [], False, 0.0, "openpyxl not installed", 0, 0)

    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    except Exception as e:
        elapsed = time.time() - t0
        return JobResult(xlsx_path, [], False, elapsed, f"failed to open workbook: {e}", 0, 0)

    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                elapsed = time.time() - t0
                return JobResult(xlsx_path, [], False, elapsed, f"sheet not found: {sheet_name}", 0, 0)
            sheetnames = [sheet_name]
        else:
            sheetnames = list(wb.sheetnames)

        out_dir = _resolve_output_dir(xlsx_path, output_dir)
        output_paths: List[Path] = []
        created = 0
        skipped = 0

        multiple = len(sheetnames) > 1 and not sheet_name
        for sn in sheetnames:
            ws = wb[sn]
            safe_sn = _sanitize_filename(sn)
            if multiple:
                csv_name = f"{xlsx_path.stem}__{safe_sn}.csv"
            else:
                csv_name = f"{xlsx_path.stem}.csv"
            csv_path = out_dir / csv_name
            output_paths.append(csv_path)

            if csv_path.exists() and not force:
                skipped += 1
                continue

            if dry_run:
                created += 1
                continue

            _write_sheet_csv(ws, csv_path)
            created += 1

        elapsed = time.time() - t0
        msg = "OK"
        return JobResult(xlsx_path, output_paths, True, elapsed, msg, created, skipped)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def print_header(root: Path, total: int, workers: int, force: bool, dry_run: bool, include_hidden: bool, output_dir: Optional[str], sheet_name: Optional[str]):
    print("xlsx-to-csv batch: Excel → CSV")
    print(f"Root: {root}")
    print(
        f"XLSX files: {total} | workers={workers} | force={force} | dry_run={dry_run} | "
        f"include_hidden={include_hidden} | output_dir={output_dir or 'same as source'} | sheet={sheet_name or 'all'}"
    )
    print("-" * 72)


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Recursively convert all .xlsx files to .csv.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--root",
        type=str,
        default="..",
        help="Root directory to scan (default: parent directory).",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=max(2, (os.cpu_count() or 4) // 2),
        help="Number of concurrent conversions (thread pool).",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Re-export even if target CSV already exists (otherwise skip).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print what would be done, without writing CSV files.",
    )
    parser.add_argument(
        "--include-hidden",
        action="store_true",
        help="Include hidden files/directories (starting with .).",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default=None,
        help="Output directory for CSV files. Defaults to same directory as source .xlsx.",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Only export the specified sheet name.",
    )
    parser.add_argument(
        "--exclude",
        nargs="*",
        default=[".git", "node_modules", ".venv", "venv", "dist", "build", "__pycache__"],
        help="Directory names to exclude anywhere in the path.",
    )

    args = parser.parse_args(argv)

    if load_workbook is None and not args.dry_run:
        print("❌ openpyxl not installed. Install it with: pip install openpyxl")
        return 2

    root = Path(args.root).expanduser().resolve()
    if not root.exists() or not root.is_dir():
        print(f"❌ Root is not a directory: {root}")
        return 2

    exclude_names = set(args.exclude or [])
    xlsx_files = find_xlsx_files(root, exclude_names, include_hidden=args.include_hidden)

    print_header(
        root=root,
        total=len(xlsx_files),
        workers=args.workers,
        force=args.force,
        dry_run=args.dry_run,
        include_hidden=args.include_hidden,
        output_dir=args.output_dir,
        sheet_name=args.sheet,
    )

    if not xlsx_files:
        print("No xlsx files found.")
        return 0

    ok_count = 0
    fail_count = 0
    skip_count = 0
    created_total = 0

    results: List[JobResult] = []
    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as ex:
        future_map = {
            ex.submit(
                convert_one,
                xlsx,
                root,
                args.output_dir,
                args.sheet,
                args.force,
                args.dry_run,
            ): xlsx
            for xlsx in xlsx_files
        }

        done_idx = 0
        total = len(future_map)

        for fut in as_completed(future_map):
            done_idx += 1
            res = fut.result()
            results.append(res)

            rel_xlsx = _human_rel(res.xlsx_path, root)

            if not res.ok:
                fail_count += 1
                print(f"[{done_idx:>4}/{total}] FAIL  {rel_xlsx}  ({res.elapsed_s:.2f}s)")
                print(f"              Reason: {res.message}")
                continue

            ok_count += 1
            created_total += res.created
            skip_count += res.skipped

            created_msg = f"created={res.created}"
            skipped_msg = f"skipped={res.skipped}"
            print(f"[{done_idx:>4}/{total}] OK    {rel_xlsx}  ({res.elapsed_s:.2f}s)  {created_msg}, {skipped_msg}")

    print("-" * 72)
    print(
        f"Done. OK={ok_count} | FAIL={fail_count} | CSV created={created_total} | "
        f"CSV skipped={skip_count} | Total XLSX={len(xlsx_files)}"
    )

    return 0 if fail_count == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
